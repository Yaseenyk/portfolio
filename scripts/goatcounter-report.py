"""Pull site analytics from the GoatCounter API into an Excel workbook.

One command turns the dashboard into a local .xlsx you can pivot, filter and
archive — pageviews, referrers, campaigns, locations, devices, plus the two
signal layers the dashboard mixes into "Pages": conversion events (evt-*) and
read-depth milestones (read:<slug>:<pct>), each broken into its own tab with
the derived numbers that matter (completion rate per post, conversions per
channel).

Usage:
    python scripts/goatcounter-report.py            # last 30 days
    python scripts/goatcounter-report.py --days 90
    python scripts/goatcounter-report.py --out my-report.xlsx

Auth:
    GOATCOUNTER_API_TOKEN in the environment or ./.env — create one at
    https://yaseen.goatcounter.com/user/api ("Read statistics" permission
    is enough). The output file is gitignored: it is analysis, not source.

API: https://www.goatcounter.com/help/api (4 req/s limit — throttled below).
"""

import argparse
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

SITE = os.environ.get("GOATCOUNTER_SITE", "yaseen")
BASE = f"https://{SITE}.goatcounter.com/api/v0"
TOKEN = os.environ.get("GOATCOUNTER_API_TOKEN", "")

# The dashboard's stat pages, one tab each.
STAT_PAGES = ["toprefs", "campaigns", "locations", "browsers", "systems", "sizes"]

HEADER_FONT = Font(bold=True)


def api(path: str, **params) -> dict:
    """GET with auth + throttle (API allows 4 req/s; 0.3s spacing is safe)."""
    time.sleep(0.3)
    res = requests.get(
        f"{BASE}{path}",
        headers={"Authorization": f"Bearer {TOKEN}"},
        params=params,
        timeout=30,
    )
    if res.status_code == 401:
        raise SystemExit(
            "401 Unauthorized — create an API token at "
            f"https://{SITE}.goatcounter.com/user/api and put it in .env as "
            "GOATCOUNTER_API_TOKEN=..."
        )
    res.raise_for_status()
    return res.json()


def add_sheet(wb: Workbook, title: str, headers: list, rows: list) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    # Fit columns to content, within reason.
    for idx, header in enumerate(headers, start=1):
        width = max(
            len(str(header)),
            *(len(str(r[idx - 1])) for r in rows[:200] if len(r) >= idx),
            8,
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 60)


def main() -> None:
    parser = argparse.ArgumentParser(description="GoatCounter → Excel report")
    parser.add_argument("--days", type=int, default=30, help="lookback window")
    parser.add_argument("--out", default="analytics-report.xlsx")
    args = parser.parse_args()

    if not TOKEN:
        raise SystemExit(
            "GOATCOUNTER_API_TOKEN not set. Create one at "
            f"https://{SITE}.goatcounter.com/user/api (Read statistics), then "
            "add GOATCOUNTER_API_TOKEN=... to .env"
        )

    end = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)
    start = end - timedelta(days=args.days)
    # GoatCounter 404s on "+00:00" offsets; it wants the trailing-Z form.
    fmt = "%Y-%m-%dT%H:%M:%SZ"
    window = {"start": start.strftime(fmt), "end": end.strftime(fmt)}
    # ASCII only — Windows consoles default to cp1252.
    print(f"window: last {args.days} days ({start.date()} -> {end.date()})")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    # ---- Overview: daily total visits -------------------------------------
    total = api("/stats/total", **window)
    daily: dict[str, int] = {}
    for stat in total.get("stats", []):
        day = str(stat.get("day", ""))[:10]
        count = stat.get("daily", 0)
        if day:
            daily[day] = daily.get(day, 0) + int(count or 0)
    add_sheet(
        wb,
        "Overview",
        ["Day", "Visits"],
        [[d, daily[d]] for d in sorted(daily)]
        + [[], ["Total", total.get("total", 0)], ["Events", total.get("total_events", 0)]],
    )
    print(f"overview: {total.get('total', 0)} visits, {total.get('total_events', 0)} events")

    # ---- Hits: pages vs conversions vs read-depth -------------------------
    hits, more, offset_guard = [], True, 0
    # /stats/hits caps limit at 100 and has no offset — one call gets the top
    # 100 paths, which covers this site; note if the API says there are more.
    res = api("/stats/hits", limit=100, **window)
    hits = res.get("hits", [])
    if res.get("more"):
        print("note: more than 100 distinct paths — tail truncated")

    pages, conversions, read_rows = [], [], {}
    for h in hits:
        path = str(h.get("path", ""))
        count = int(h.get("count", 0) or 0)
        title = str(h.get("title", ""))
        m = re.match(r"^read:(.+):(25|50|75|100)$", path)
        if m:
            slug, milestone = m.group(1), int(m.group(2))
            read_rows.setdefault(slug, {})[milestone] = count
        elif path.startswith("evt-") or h.get("event"):
            conversions.append([path.removeprefix("evt-"), count])
        else:
            pages.append([path, title, count])

    pages.sort(key=lambda r: -r[2])
    conversions.sort(key=lambda r: -r[1])
    add_sheet(wb, "Pages", ["Path", "Title", "Visits"], pages)
    add_sheet(wb, "Conversions", ["Event", "Count"], conversions)

    depth_table = []
    for slug, m in sorted(read_rows.items(), key=lambda kv: -kv[1].get(25, 0)):
        p25, p50, p75, p100 = (m.get(k, 0) for k in (25, 50, 75, 100))
        completion = f"{(p100 / p25 * 100):.0f}%" if p25 else ""
        depth_table.append([slug, p25, p50, p75, p100, completion])
    add_sheet(
        wb,
        "Read depth",
        ["Post", "25%", "50%", "75%", "100%", "Completion (100÷25)"],
        depth_table,
    )
    print(
        f"hits: {len(pages)} pages, {len(conversions)} conversion events, "
        f"{len(depth_table)} posts with read-depth"
    )

    # ---- Referrers, campaigns, locations, devices -------------------------
    titles = {
        "toprefs": "Referrers",
        "campaigns": "Campaigns",
        "locations": "Locations",
        "browsers": "Browsers",
        "systems": "Systems",
        "sizes": "Screen sizes",
    }
    for page in STAT_PAGES:
        stats, offset = [], 0
        while True:
            res = api(f"/stats/{page}", limit=100, offset=offset, **window)
            stats.extend(res.get("stats", []))
            if not res.get("more") or offset >= 400:
                break
            offset += 100
        rows = [[s.get("name", s.get("id", "")), int(s.get("count", 0) or 0)] for s in stats]
        rows.sort(key=lambda r: -r[1])
        add_sheet(wb, titles[page], [titles[page].rstrip("s") or "Name", "Visits"], rows)
        print(f"{page}: {len(rows)} rows")

    out = Path(args.out)
    wb.save(out)
    print(f"\nsaved: {out.resolve()}")


if __name__ == "__main__":
    main()
